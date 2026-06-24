"""Excel report generation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .constants import (
    AUTO_HOLD_WORKSHEET_NAMES,
    CORE_CANDIDATES_WORKSHEET_NAMES,
    MANUAL_REVIEW_SHEET_NAME,
    WORKSHEET_NAMES,
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
STATUS_FILLS = {
    "PASS": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "PRICE_READY": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "PASS_LOGISTICS": PatternFill(fill_type="solid", fgColor="EAF4E3"),
    "REVIEW": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "PRICE_REVIEW": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "REVIEW_LOGISTICS": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "HOLD_LOGISTICS": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "OUT_OF_STOCK": PatternFill(fill_type="solid", fgColor="E7E6E6"),
    "ORPHAN": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "OTHER_SUPPLIER": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "OTHER_SUPPLIER_CONFIRMED": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "CONFLICT": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "CRITICAL": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "UNKNOWN_SUPPLIER": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "BENZARA_ORPHAN_SUSPECTED": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "AUTO_PASS": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "AUTO_HOLD_LOGISTICS": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "AUTO_HOLD_OUT_OF_STOCK": PatternFill(fill_type="solid", fgColor="E7E6E6"),
    "AUTO_ARCHIVE_ORPHAN_CANDIDATE": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "KEEP_OUTSIDE_BENZARA_FLOW": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "MANUAL_REVIEW_HIGH": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "MANUAL_REVIEW_MEDIUM": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "MANUAL_REVIEW_LOW": PatternFill(fill_type="solid", fgColor="EAF4E3"),
    "HIGH": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "MEDIUM": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "LOW": PatternFill(fill_type="solid", fgColor="EAF4E3"),
    "HIGH_PRIORITY": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "MEDIUM_PRIORITY": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "LOW_PRIORITY": PatternFill(fill_type="solid", fgColor="EAF4E3"),
}
CURRENCY_COLUMNS = {
    "Current Regular Price",
    "Current Sale Price",
    "Benzara Regular Price",
    "Recommended Price Update",
    "Current Price",
    "Benzara Price",
    "Price",
    "Regular price",
    "Sale price",
}
DECIMAL_COLUMNS = {
    "Weight lb",
    "Length in",
    "Width in",
    "Height in",
    "Volume in3",
    "Dim Weight lb",
    "Length + Girth in",
    "actual_weight_lb",
    "length_in",
    "width_in",
    "height_in",
    "volume_in3",
    "dim_weight_lb",
    "girth_in",
    "length_plus_girth_in",
    "billable_weight_lb",
    "longest_side_in",
    "Weight",
    "Priority Score",
    "Length",
    "Width",
    "Height",
    "Volume",
    "Dim Weight",
    "Length + Girth",
}
INTEGER_COLUMNS = {
    "WooCommerce ID",
    "Benzara Stock Qty",
    "Current Stock Qty",
    "Woo Candidate Count",
    "Stock Qty",
    "Sort Score",
}
STATUS_COLUMNS = {
    "Catalog Decision",
    "Logistics Status",
    "Commercial Status",
    "Price Update Status",
    "Data Quality Status",
    "Supplier Classification",
    "Priority",
    "Review Type",
    "Review Priority",
    "Review Batch",
    "Source Bucket",
    "Recommended Operational Action",
}
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


def build_workbook(sheet_names: tuple[str, ...]) -> Workbook:
    workbook = Workbook()
    workbook.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        workbook.create_sheet(title=name)
    return workbook


def _autosize_and_filter(worksheet) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max((len(value) for value in values), default=10) + 2, 60)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def _apply_formats(worksheet) -> None:
    if worksheet.max_row < 2:
        return
    header_map = {cell.column: cell.value for cell in worksheet[1]}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = header_map.get(cell.column)
            if header in CURRENCY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"
            elif header in DECIMAL_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
            elif header in INTEGER_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "0"
            if header in STATUS_COLUMNS:
                fill = STATUS_FILLS.get(str(cell.value or ""))
                if fill is not None:
                    cell.fill = fill


def _add_table(worksheet, sheet_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    if worksheet[1][0].value == "Message":
        return
    table_name = f"T_{''.join(character for character in sheet_name if character.isalnum())[:24]}"
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TABLE_STYLE
    worksheet.add_table(table)


def _normalize_cell_value(value):
    if isinstance(value, (list, tuple, set)):
        return ";".join(str(item) for item in value)
    if isinstance(value, dict):
        return ";".join(f"{key}={item}" for key, item in value.items())
    return value


def _write_table(worksheet, rows: list[dict], sheet_name: str) -> None:
    if not rows:
        worksheet.append(["Message"])
        worksheet.append(["No rows"])
        _style_header(worksheet[1])
        _autosize_and_filter(worksheet)
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    _style_header(worksheet[1])
    for row in rows:
        worksheet.append([_normalize_cell_value(row.get(header)) for header in headers])
    _apply_formats(worksheet)
    _add_table(worksheet, sheet_name)
    _autosize_and_filter(worksheet)


def write_named_workbook(output_path: str | Path, sheet_names: tuple[str, ...], sections: dict[str, list[dict]]) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(sheet_names)
    for sheet_name in sheet_names:
        _write_table(workbook[sheet_name], sections.get(sheet_name, []), sheet_name)
    workbook.save(path)
    return path


def write_workbook(output_path: str | Path, report_sections: dict[str, list[dict]], summary_rows: list[dict], rules_rows: list[dict]) -> Path:
    sections = dict(report_sections)
    sections["SUMMARY"] = summary_rows
    sections["RULES_AND_CONFIG"] = rules_rows
    return write_named_workbook(output_path, WORKSHEET_NAMES, sections)


def write_manual_review_workbook(output_path: str | Path, manual_review_rows: list[dict]) -> Path:
    return write_named_workbook(output_path, (MANUAL_REVIEW_SHEET_NAME,), {MANUAL_REVIEW_SHEET_NAME: manual_review_rows})


def write_auto_hold_workbook(output_path: str | Path, summary_rows: list[dict], sections: dict[str, list[dict]]) -> Path:
    workbook_sections = dict(sections)
    workbook_sections["SUMMARY"] = summary_rows
    return write_named_workbook(output_path, AUTO_HOLD_WORKSHEET_NAMES, workbook_sections)


def write_core_candidates_workbook(output_path: str | Path, core_candidate_rows: list[dict]) -> Path:
    return write_named_workbook(output_path, CORE_CANDIDATES_WORKSHEET_NAMES, {"CORE_CANDIDATES": core_candidate_rows})
